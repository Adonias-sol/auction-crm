import re
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db import transaction
from django.utils import timezone
from rest_framework import viewsets, status, mixins
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import ImportBatch, Winner, Invoice, InvoiceLot, FeeConfig
from .serializers import ImportBatchSerializer, InvoiceListSerializer
from .permissions import has_permission

# Replace "from .views import StandardPagination" with:
from .pagination import StandardPagination


REQUIRED_COLUMNS = [
    'Lot No', 'Auction', 'Initial Price', 'Name', 'Phone number',
    'Amount', 'Submitted At', 'CPO Amount', 'CPO Bank', 'Status',
]


def parse_submitted_at(raw):
    """
    bid_data_report.xlsx's 'Submitted At' column looks like
    'Nov. 20, 2025, 1:13 p.m.' — not ISO, and %b/%p in strptime don't
    accept the periods after the month or in "p.m." directly, so this
    normalizes those before parsing.
    """
    if raw in (None, ''):
        return None
    text = str(raw).strip()
    text = re.sub(r'^([A-Za-z]{3})\.', r'\1', text)                       # "Nov." -> "Nov"
    text = re.sub(r'([ap])\.m\.$', lambda m: m.group(1).upper() + 'M',    # "p.m." -> "PM"
                  text, flags=re.IGNORECASE)
    try:
        dt = datetime.strptime(text, '%b %d, %Y, %I:%M %p')
    except ValueError:
        return None
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt)
    return dt


def to_decimal(value):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


def parse_bid_report(file_obj):
    """
    Reads the header row to map column name -> position, so column order
    in the file doesn't matter as long as the names match. Returns
    (rows, error_message).
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = wb.active
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    missing = [c for c in REQUIRED_COLUMNS if c not in header_row]
    if missing:
        return None, f"Missing expected column(s): {', '.join(missing)}"

    idx = {name: header_row.index(name) for name in REQUIRED_COLUMNS}

    rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[idx['Name']] is None and row[idx['Phone number']] is None:
            continue  # skip fully blank trailing rows
        rows.append({
            'rowNumber': row_number,
            'lotNumber': row[idx['Lot No']],
            'auctionName': row[idx['Auction']],
            'initialPrice': row[idx['Initial Price']],
            'bidderName': row[idx['Name']],
            'winnerPhone': row[idx['Phone number']],
            'winningAmount': row[idx['Amount']],
            'submittedAt': row[idx['Submitted At']],
            'cpoAmount': row[idx['CPO Amount']],
            'cpoBank': row[idx['CPO Bank']],
            'status': row[idx['Status']],
        })
    return rows, None


class ImportBatchPreviewView(APIView):
    """
    POST /api/import-batches/preview/ — parses the file, groups Winner rows
    by phone number, computes default fees. Saves NOTHING — matches the
    frontend's "Preview — nothing saved yet" screen. Staff review/edit fee
    percentages client-side; confirm/ creates records from that (possibly
    edited) grouped data, not by re-parsing the file a second time.
    """
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file_obj = request.FILES.get('file')
        company_name = request.data.get('companyName')
        auction_date = request.data.get('auctionDate')

        if not file_obj:
            return Response({'file': ['This field is required.']}, status=status.HTTP_400_BAD_REQUEST)
        if not company_name or not auction_date:
            return Response({'error': 'companyName and auctionDate are required'}, status=status.HTTP_400_BAD_REQUEST)

        rows, header_error = parse_bid_report(file_obj)
        if header_error:
            return Response({'error': header_error}, status=status.HTTP_400_BAD_REQUEST)

        default_fee_pct = FeeConfig.get_active_percentage()
        groups = {}
        flagged_rows = []
        valid_lot_count = 0

        for row in rows:
            if row['status'] != 'Winner':
                continue  # Submitted rows are dropped entirely, per the import rules

            issues = []
            phone = row['winnerPhone']
            phone = str(int(phone)) if isinstance(phone, float) else (str(phone).strip() if phone else '')
            name = (row['bidderName'] or '').strip()
            amount = to_decimal(row['winningAmount'])

            if not name:
                issues.append('missing bidder name')
            if not phone:
                issues.append('missing phone number')
            if amount is None:
                issues.append('missing or invalid winning amount')

            if issues:
                flagged_rows.append({'rowNumber': row['rowNumber'], 'data': row, 'issues': issues})
                continue

            lot_fee = (amount * default_fee_pct / Decimal('100')).quantize(Decimal('0.01'))
            lot = {
                'lotNumber': row['lotNumber'] or '',
                'auctionName': row['auctionName'] or '',
                'initialPrice': str(to_decimal(row['initialPrice']) or ''),
                'winningAmount': str(amount),
                'cpoAmount': str(to_decimal(row['cpoAmount']) or ''),
                'cpoBank': row['cpoBank'] or '',
                'feePercentage': str(default_fee_pct),
                'lotFee': str(lot_fee),
                'submittedAt': row['submittedAt'].isoformat() if hasattr(row['submittedAt'], 'isoformat') else row['submittedAt'],
            }

            if phone not in groups:
                groups[phone] = {
                    'bidderName': name,
                    'winnerPhone': phone,
                    'companyName': '',
                    'feePercentage': str(default_fee_pct),
                    'lots': [],
                }
            groups[phone]['lots'].append(lot)
            valid_lot_count += 1

        for group in groups.values():
            group['totalFee'] = str(sum(Decimal(l['lotFee']) for l in group['lots']))

        return Response({
            'groupedWinners': list(groups.values()),
            'flaggedRows': flagged_rows,
            'totalWinners': len(groups),
            'totalLots': valid_lot_count,
            'flaggedCount': len(flagged_rows),
        })


class ImportBatchConfirmView(APIView):
    """
    POST /api/import-batches/confirm/ — creates ImportBatch + Winner +
    Invoice + InvoiceLot records from the grouped data the preview screen
    returned (with whatever fee% edits staff made in the browser).

    DESIGN DECISION worth knowing: Winner has single winningAmount/
    initialPrice/cpoAmount fields, but one winner can win several lots.
    Resolved by making each Winner record represent that bidder's TOTAL
    across all their lots in this one batch (sums), while every
    individual lot's own numbers live on InvoiceLot, which is already
    fully itemized per lot. This is what keeps Invoice.winner a single
    clean ForeignKey instead of needing a many-to-many relationship —
    this wasn't explicit anywhere in the original model spec, so flagging
    it here rather than leaving it implicit in the code.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not has_permission(request.user, 'generate_invoice'):
            # Confirming an import creates invoices — same permission bar
            # as generating one PDF (admin + auction_manager).
            return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)

        file_name = request.data.get('fileName', '')
        company_name = request.data.get('companyName')
        auction_date = request.data.get('auctionDate')
        batch_name = request.data.get('batchName', '')
        grouped_winners = request.data.get('groupedWinners', [])
        invalid_records = int(request.data.get('invalidRecords', 0) or 0)
        # Optional: how many days out the due date defaults to. 14 is a
        # placeholder — swap for whatever finance's actual payment terms
        # are, or accept it per-batch from the request if it varies.
        due_date_input = request.data.get('dueDate')  # e.g. "2026-09-03", from the preview screen
        due_in_days = int(request.data.get('dueInDays', 14))
        invoice_due_date = due_date_input or (timezone.localdate() + timedelta(days=due_in_days))

        if not company_name or not auction_date:
            return Response({'error': 'companyName and auctionDate are required'}, status=status.HTTP_400_BAD_REQUEST)
        if not grouped_winners:
            return Response({'error': 'groupedWinners must contain at least one winner'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            total_lots = sum(len(g['lots']) for g in grouped_winners)
            batch = ImportBatch.objects.create(
                fileName=file_name,
                batchName=batch_name,
                companyName=company_name,
                auctionDate=auction_date,
                status='confirmed',
                totalRecords=total_lots + invalid_records,
                validRecords=total_lots,
                invalidRecords=invalid_records,
                importedBy=request.user,
            )

            created_invoice_ids = []
            for group in grouped_winners:
                lots = group['lots']
                total_winning = sum(Decimal(l['winningAmount']) for l in lots)
                total_cpo = sum((Decimal(l['cpoAmount']) if l.get('cpoAmount') else Decimal('0')) for l in lots)
                total_initial = sum((Decimal(l['initialPrice']) if l.get('initialPrice') else Decimal('0')) for l in lots)
                parsed_submitted_dates = [
                    d for d in (parse_submitted_at(l.get('submittedAt')) for l in lots)
                    if d is not None
                ]
                earliest_submitted = min(parsed_submitted_dates, default=None)

                winner = Winner.objects.create(
                    bidderName=group['bidderName'],
                    winnerPhone=group['winnerPhone'],
                    companyName=group.get('companyName', ''),
                    winningAmount=total_winning,
                    cpoAmount=total_cpo or None,
                    initialPrice=total_initial or None,
                    submittedAt=earliest_submitted,
                    importBatch=batch,
                )

                invoice = Invoice.objects.create(
                    winner=winner,
                    importBatch=batch,
                    invoiceNumber=self._next_invoice_number(),
                    invoiceDate=timezone.localdate(),
                    dueDate=invoice_due_date,
                    status='invoice_generated',
                )

                for lot in lots:
                    raw_initial = lot.get('initialPrice')
                    raw_cpo = lot.get('cpoAmount')
                    raw_fee_pct = group.get('feePercentage', lot.get('feePercentage'))
                    InvoiceLot.objects.create(
                        invoice=invoice,
                        lotNumber=lot['lotNumber'],
                        auctionName=lot['auctionName'],
                        initialPrice=Decimal(raw_initial) if raw_initial else None,
                        winningAmount=Decimal(lot['winningAmount']),
                        cpoAmount=Decimal(raw_cpo) if raw_cpo else None,
                        cpoBank=lot.get('cpoBank', ''),
                        feePercentage=Decimal(raw_fee_pct),
                        submittedAt=parse_submitted_at(lot.get('submittedAt')),
                        # lotFee isn't passed — InvoiceLot.save() computes it.
                    )
                created_invoice_ids.append(invoice.id)

        return Response({
            'importBatch': ImportBatchSerializer(batch).data,
            'invoiceIds': created_invoice_ids,
        }, status=status.HTTP_201_CREATED)

    @staticmethod
    def _next_invoice_number():
        # Simple incrementing scheme: INV-<year>-<seq>. Fine at this scale;
        # swap to select_for_update() if concurrent imports ever race.
        year = timezone.localdate().year
        count = Invoice.objects.filter(invoiceNumber__startswith=f'INV-{year}-').count()
        return f'INV-{year}-{count + 1:03d}'


class ImportBatchViewSet(mixins.DestroyModelMixin, viewsets.ReadOnlyModelViewSet):
    """GET /api/import-batches/ and /api/import-batches/{id}/ — read-only, confirm/ is what creates them."""
    queryset = ImportBatch.objects.select_related('importedBy').order_by('-uploadDate')
    serializer_class = ImportBatchSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    @action(detail=True, methods=['get'])
    def invoices(self, request, pk=None):
        batch = self.get_object()
        qs = Invoice.objects.select_related('winner').filter(importBatch=batch)
        page = self.paginate_queryset(qs)
        serializer = InvoiceListSerializer(page if page is not None else qs, many=True)
        if page is not None:
            return self.get_paginated_response(serializer.data)
        return Response(serializer.data)
    
    def perform_destroy(self, instance):
        if not has_permission(self.request.user, 'delete_records'):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only administrators can delete import batches.')
        with transaction.atomic():
            # Invoice.delete() cascades to InvoiceLot/Payment/Attachment/AuditLog already
            Invoice.objects.filter(importBatch=instance).delete()
            Winner.objects.filter(importBatch=instance).delete()
            instance.delete()